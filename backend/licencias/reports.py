"""
Generación de reportes Excel y PDF.
"""
import io
from datetime import datetime
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from .permissions import PuedeVerReportesYPersonal
from .models import Licencia, Usuario


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportesYPersonal])
def exportar_excel(request):
    """Exportar reporte de licencias a Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Licencias Médicas'

    # Filtros
    fecha_desde = request.query_params.get('fecha_desde')
    fecha_hasta = request.query_params.get('fecha_hasta')
    tipo = request.query_params.get('tipo')
    estado = request.query_params.get('estado')

    licencias = Licencia.objects.select_related(
        'usuario', 'usuario__jerarquia', 'usuario__dependencia',
        'usuario__dependencia__ciudad'
    ).all()

    if fecha_desde:
        licencias = licencias.filter(fecha_inicio__gte=fecha_desde)
    if fecha_hasta:
        licencias = licencias.filter(fecha_inicio__lte=fecha_hasta)
    if tipo:
        licencias = licencias.filter(tipo=tipo)
    if estado:
        licencias = licencias.filter(estado=estado)

    # Estilos
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = 'REPORTE DE LICENCIAS MÉDICAS'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1E40AF')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(name='Arial', italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        'Personal', 'Jerarquía', 'Dependencia', 'Ciudad',
        'Tipo Licencia', 'Estado', 'Fecha Inicio', 'Fecha Fin', 'Días',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    for row, lic in enumerate(licencias, 5):
        datos = [
            lic.usuario.nombre_completo,
            lic.usuario.jerarquia.nombre if lic.usuario.jerarquia else '',
            lic.usuario.dependencia.nombre if lic.usuario.dependencia else '',
            lic.usuario.dependencia.ciudad.nombre if lic.usuario.dependencia and lic.usuario.dependencia.ciudad else '',
            lic.get_tipo_display(),
            lic.get_estado_display(),
            lic.fecha_inicio.strftime('%d/%m/%Y'),
            lic.fecha_fin.strftime('%d/%m/%Y'),
            lic.dias_licencia,
        ]
        for col, dato in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=dato)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # Ajustar anchos
    anchos = [30, 20, 30, 15, 20, 15, 15, 15, 8]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[chr(64 + col)].width = ancho

    # Generar respuesta
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="licencias_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated, PuedeVerReportesYPersonal])
def exportar_pdf(request):
    """Exportar reporte de licencias a PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    # Filtros
    fecha_desde = request.query_params.get('fecha_desde')
    fecha_hasta = request.query_params.get('fecha_hasta')
    tipo = request.query_params.get('tipo')
    estado = request.query_params.get('estado')

    licencias = Licencia.objects.select_related(
        'usuario', 'usuario__jerarquia', 'usuario__dependencia',
        'usuario__dependencia__ciudad'
    ).all()

    if fecha_desde:
        licencias = licencias.filter(fecha_inicio__gte=fecha_desde)
    if fecha_hasta:
        licencias = licencias.filter(fecha_inicio__lte=fecha_hasta)
    if tipo:
        licencias = licencias.filter(tipo=tipo)
    if estado:
        licencias = licencias.filter(estado=estado)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor('#1E40AF'),
        spaceAfter=6,
    )
    elements.append(Paragraph('REPORTE DE LICENCIAS MÉDICAS', title_style))

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20,
    )
    elements.append(Paragraph(
        f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        subtitle_style
    ))

    # Tabla
    data = [['Personal', 'Jerarquía', 'Dependencia', 'Ciudad', 'Tipo', 'Estado', 'Inicio', 'Fin', 'Días']]

    for lic in licencias:
        data.append([
            lic.usuario.nombre_completo,
            lic.usuario.jerarquia.nombre if lic.usuario.jerarquia else '',
            lic.usuario.dependencia.nombre if lic.usuario.dependencia else '',
            lic.usuario.dependencia.ciudad.nombre if lic.usuario.dependencia and lic.usuario.dependencia.ciudad else '',
            lic.get_tipo_display(),
            lic.get_estado_display(),
            lic.fecha_inicio.strftime('%d/%m/%Y'),
            lic.fecha_fin.strftime('%d/%m/%Y'),
            str(lic.dias_licencia),
        ])

    col_widths = [5 * cm, 3.5 * cm, 4.5 * cm, 2.5 * cm, 3 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 1.5 * cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (-1, 1), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4FF')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f'Total de licencias: {len(data) - 1}',
        styles['Normal']
    ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="licencias_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response
